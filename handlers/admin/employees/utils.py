import io
import json
import logging
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from db import get_connection
from db.users import get_user
from db.profile import (
    get_employee_full_info,
    get_taxi_summary,
    get_salary_history,
    get_current_salary_rate,
)
from db.checklist import get_items_for_location_and_day

from utils.time_utils import today_msk_str

logger = logging.getLogger(__name__)


# =========================================================
# DESIGN SYSTEM
# =========================================================
COLOR_DARK = "1D1D1F"
COLOR_LIGHT = "F5F5F7"
COLOR_BLUE = "0071E3"
COLOR_GREEN = "34C759"
COLOR_ORANGE = "FF9500"
COLOR_WHITE = "FFFFFF"
COLOR_GRAY = "6E6E73"

FONT_NAME = "Arial"

THIN_BORDER = Border(
    left=Side(style="thin", color="D2D2D7"),
    right=Side(style="thin", color="D2D2D7"),
    top=Side(style="thin", color="D2D2D7"),
    bottom=Side(style="thin", color="D2D2D7"),
)


def _safe(value) -> str:
    if value is None:
        return "—"
    return str(value)


def _money(value) -> float:
    try:
        return round(float(value or 0), 2)
    except Exception:
        return 0.0


def _num(value, digits: int = 1) -> float:
    try:
        return round(float(value or 0), digits)
    except Exception:
        return 0.0


def _safe_percent(value) -> str:
    if value is None:
        return "—"
    try:
        return f"{round(float(value), 1)}%"
    except Exception:
        return "—"


def _truncate(value, limit: int) -> str:
    text = _safe(value)
    if len(text) <= limit:
        return text
    return text[:limit] + "…"


def _style_title(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name=FONT_NAME, size=16, bold=True, color=COLOR_DARK)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 28
    return cell


def _style_subtitle(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name=FONT_NAME, size=11, color=COLOR_GRAY)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 18
    return cell


def _style_section(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_BLUE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22
    return cell


def _style_header_row(ws, row: int, headers: list[str]):
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=header)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill("solid", fgColor=COLOR_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def _style_data_row(ws, row: int, values: list, zebra: bool):
    fill = PatternFill("solid", fgColor=COLOR_LIGHT) if zebra else None

    for idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=idx, value=value)
        cell.font = Font(name=FONT_NAME, size=10, color=COLOR_DARK)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

        if fill:
            cell.fill = fill

    ws.row_dimensions[row].height = 20


def _style_metric_row(ws, row: int, label: str, value, zebra: bool = False):
    fill = PatternFill("solid", fgColor=COLOR_LIGHT) if zebra else None

    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_GRAY)
    label_cell.alignment = Alignment(vertical="center")
    label_cell.border = THIN_BORDER

    value_cell = ws.cell(row=row, column=2, value=value)
    value_cell.font = Font(name=FONT_NAME, size=10, color=COLOR_DARK)
    value_cell.alignment = Alignment(vertical="center", wrap_text=True)
    value_cell.border = THIN_BORDER

    if fill:
        label_cell.fill = fill
        value_cell.fill = fill

    ws.row_dimensions[row].height = 20


def _auto_width(ws, min_width: int = 12, max_width: int = 40):
    for column_cells in ws.columns:
        if not column_cells:
            continue

        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        adjusted = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted


# =========================================================
# BASE DATA HELPERS
# =========================================================
def _period_range(period_days: int = 30) -> tuple[str, str]:
    date_to = today_msk_str()
    date_from = (datetime.now() - timedelta(days=period_days)).strftime("%Y-%m-%d")
    return date_from, date_to


def _parse_date(value: str):
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _days_since(date_str: str) -> int | None:
    d = _parse_date(date_str)
    if not d:
        return None
    return (datetime.now().date() - d).days


def _date_range(date_from: str, date_to: str):
    start = _parse_date(date_from)
    end = _parse_date(date_to)

    if not start:
        return

    if not end:
        end = start

    current = start

    while current <= end:
        yield current.strftime("%Y-%m-%d")
        current += timedelta(days=1)


def _normalize_location(value) -> str:
    if not value:
        return "Не указано"

    val = str(value).strip().lower()

    if val in {"bar", "бар"} or "бар" in val:
        return "Бар"

    if val in {"kitchen", "кухня"} or "кух" in val:
        return "Кухня"

    return str(value).strip().capitalize() or "Не указано"


def _zero_daily() -> dict:
    return {
        "shifts": 0,
        "hours": 0.0,
        "earned": 0.0,
        "taxi": 0.0,
        "tasks": 0,
        "reports": 0,
    }


def _has_photo_progress(row: dict) -> bool:
    if row.get("photo_file_id"):
        return True

    raw = row.get("photo_file_ids")

    if not raw:
        return False

    try:
        parsed = json.loads(raw)
    except Exception:
        return bool(str(raw).strip())

    if isinstance(parsed, list):
        for entry in parsed:
            if isinstance(entry, str) and entry.strip():
                return True
            if isinstance(entry, dict) and entry.get("file_id"):
                return True
        return False

    if isinstance(parsed, dict):
        return bool(parsed.get("file_id"))

    return bool(parsed)


def get_employee_shifts(user_id: int, date_from: str, date_to: str) -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                s.date,
                s.start_time,
                s.active,
                st.name AS shift_name,
                st.location,
                st.duration
            FROM shifts s
            LEFT JOIN shift_types st ON s.shift_type_id = st.id
            WHERE s.user_id = ?
              AND s.date >= ?
              AND s.date <= ?
            ORDER BY s.date DESC, s.start_time DESC
            """,
            (user_id, date_from, date_to),
        ).fetchall()

    return [dict(row) for row in rows]


def get_employee_reports(user_id: int, date_from: str, date_to: str) -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                date,
                report_type,
                full_text,
                created_at,
                updated_at
            FROM shift_reports
            WHERE author_id = ?
              AND date >= ?
              AND date <= ?
            ORDER BY date DESC
            """,
            (user_id, date_from, date_to),
        ).fetchall()

    return [dict(row) for row in rows]


def get_employee_checklist_activity(user_id: int, date_from: str, date_to: str) -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                p.date,
                p.completed,
                p.completed_at,
                p.item_id,
                p.photo_file_id,
                p.photo_file_ids,
                i.text AS item_text,
                i.location,
                i.category,
                i.requires_photo
            FROM checklist_shared_progress p
            LEFT JOIN checklist_items i ON p.item_id = i.id
            WHERE p.completed_by = ?
              AND p.date >= ?
              AND p.date <= ?
              AND COALESCE(p.completed, 0) = 1
            ORDER BY p.date DESC
            """,
            (user_id, date_from, date_to),
        ).fetchall()

    result = []

    for row in rows:
        item = dict(row)
        item["has_photo"] = _has_photo_progress(item)
        result.append(item)

    return result


def get_taxi_expenses_full(user_id: int, date_from: str, date_to: str) -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                date,
                amount,
                photo_file_ids
            FROM taxi_expenses
            WHERE user_id = ?
              AND date >= ?
              AND date <= ?
            ORDER BY date DESC
            """,
            (user_id, date_from, date_to),
        ).fetchall()

    result = []

    for row in rows:
        item = dict(row)
        photos = []
        raw = item.get("photo_file_ids")

        if raw:
            try:
                parsed = json.loads(raw)

                if isinstance(parsed, list):
                    for entry in parsed:
                        if isinstance(entry, str):
                            photos.append(entry)
                        elif isinstance(entry, dict) and entry.get("file_id"):
                            photos.append(entry["file_id"])
            except Exception:
                pass

        item["photos"] = photos
        result.append(item)

    return result


# =========================================================
# ANALYTICS HELPERS
# =========================================================
def _get_rate_for_date(user_id: int, date: str, fallback_rate: float, cache: dict) -> float:
    key = (user_id, date)

    if key in cache:
        return cache[key]

    try:
        rate = get_current_salary_rate(user_id, date)
    except Exception:
        rate = None

    if rate is None:
        rate = fallback_rate or 0

    cache[key] = float(rate or 0)
    return cache[key]


def _max_consecutive_days(date_strings) -> int:
    dates = sorted({d for d in (_parse_date(x) for x in date_strings) if d})

    if not dates:
        return 0

    max_count = 1
    current = 1
    prev = dates[0]

    for d in dates[1:]:
        if d == prev + timedelta(days=1):
            current += 1
        else:
            current = 1

        if current > max_count:
            max_count = current

        prev = d

    return max_count


def _build_expected_tasks(shifts: list[dict]) -> set:
    expected = set()

    for shift in shifts:
        location = shift.get("location")
        date = shift.get("date")

        if not location or not date:
            continue

        try:
            items = get_items_for_location_and_day(location, date)
        except Exception as e:
            logger.warning("Не удалось получить ожидаемые чек-листы для %s %s: %s", location, date, e)
            items = []

        for item in items:
            item_id = item.get("id")
            if item_id:
                expected.add((date, location, item_id))

    return expected


def _employee_metrics(
    user: dict,
    date_from: str,
    date_to: str,
    period_days: int,
    rate_cache: dict | None = None,
) -> dict:
    tg_id = user.get("tg_id") or user.get("id")
    period_days = max(int(period_days or 1), 1)

    if rate_cache is None:
        rate_cache = {}

    info = get_employee_full_info(tg_id) or {}
    fallback_rate = float(info.get("current_rate") or 0)

    shifts_raw = get_employee_shifts(tg_id, date_from, date_to)
    reports = get_employee_reports(tg_id, date_from, date_to)
    activity = get_employee_checklist_activity(tg_id, date_from, date_to)
    taxi_summary = get_taxi_summary(tg_id, date_from, date_to)
    taxi_expenses = get_taxi_expenses_full(tg_id, date_from, date_to)

    shift_rows = []
    daily = {}

    total_hours = 0.0
    total_earned = 0.0

    for shift in shifts_raw:
        row = dict(shift)

        hours = _num((row.get("duration") or 0) / 60, 2)
        rate = _get_rate_for_date(tg_id, row.get("date"), fallback_rate, rate_cache)
        earned = _num(hours * rate, 2)

        row["hours"] = hours
        row["rate"] = rate
        row["earned"] = earned

        shift_rows.append(row)

        total_hours += hours
        total_earned += earned

        d = row.get("date")

        if d:
            daily.setdefault(d, _zero_daily())
            daily[d]["shifts"] += 1
            daily[d]["hours"] += hours
            daily[d]["earned"] += earned

    expected_set = _build_expected_tasks(shift_rows)

    completed_expected = 0
    tasks_with_photo = 0
    required_photo_missing = 0

    for act in activity:
        if act.get("has_photo"):
            tasks_with_photo += 1

        if act.get("requires_photo") and not act.get("has_photo"):
            required_photo_missing += 1

        key = (act.get("date"), act.get("location"), act.get("item_id"))

        if key in expected_set:
            completed_expected += 1

        d = act.get("date")

        if d:
            daily.setdefault(d, _zero_daily())
            daily[d]["tasks"] += 1

    for expense in taxi_expenses:
        d = expense.get("date")
        amount = _money(expense.get("amount"))

        if d:
            daily.setdefault(d, _zero_daily())
            daily[d]["taxi"] += amount

    for report in reports:
        d = report.get("date")

        if d:
            daily.setdefault(d, _zero_daily())
            daily[d]["reports"] += 1

    shift_dates = {s.get("date") for s in shift_rows if s.get("date")}
    report_dates = {r.get("date") for r in reports if r.get("date")}
    covered_report_dates = shift_dates & report_dates

    report_coverage_percent = None
    if shift_dates:
        report_coverage_percent = _num(len(covered_report_dates) / len(shift_dates) * 100, 1)

    expected_tasks = len(expected_set)
    checklist_percent = None
    if expected_tasks:
        checklist_percent = _num(completed_expected / expected_tasks * 100, 1)

    tasks_completed = len(activity)
    photo_percent = None
    if tasks_completed:
        photo_percent = _num(tasks_with_photo / tasks_completed * 100, 1)

    total_taxi = _money(taxi_summary.get("total"))
    total_cost = _money(total_earned + total_taxi)

    cost_per_hour = 0.0
    if total_hours:
        cost_per_hour = _num(total_cost / total_hours, 2)

    average_rate = fallback_rate
    if total_hours:
        average_rate = _num(total_earned / total_hours, 2)

    taxi_share_percent = 0.0
    if total_cost:
        taxi_share_percent = _num(total_taxi / total_cost * 100, 2)

    hours_per_week = _num(total_hours * 7 / period_days, 1)

    last_shift_date = max(shift_dates) if shift_dates else None
    first_shift_date = min(shift_dates) if shift_dates else None
    days_since_last_shift = _days_since(last_shift_date) if last_shift_date else None
    max_consecutive_days = _max_consecutive_days(shift_dates)

    reports_opening = sum(1 for r in reports if r.get("report_type") == "opening")
    reports_closing = sum(1 for r in reports if r.get("report_type") != "opening")

    return {
        "id": tg_id,
        "name": user.get("full_name") or user.get("first_name") or f"ID {tg_id}",
        "position": info.get("position") or user.get("position"),
        "status": info.get("status") or user.get("status"),
        "is_active": bool(user.get("is_active", 1)),
        "current_rate": fallback_rate,
        "shifts": shift_rows,
        "reports": reports,
        "activity": activity,
        "taxi_expenses": taxi_expenses,
        "daily": daily,

        "shift_count": len(shift_rows),
        "total_hours": _num(total_hours, 1),
        "earned_total": _money(total_earned),
        "taxi_total": total_taxi,
        "total_cost": total_cost,
        "cost_per_hour": cost_per_hour,
        "average_rate": average_rate,
        "taxi_share_percent": taxi_share_percent,

        "reports_count": len(reports),
        "reports_opening": reports_opening,
        "reports_closing": reports_closing,
        "shift_dates_count": len(shift_dates),
        "covered_report_dates_count": len(covered_report_dates),
        "report_coverage_percent": report_coverage_percent,

        "tasks_completed": tasks_completed,
        "expected_tasks": expected_tasks,
        "completed_expected_tasks": completed_expected,
        "checklist_percent": checklist_percent,
        "tasks_with_photo": tasks_with_photo,
        "photo_percent": photo_percent,
        "required_photo_missing": required_photo_missing,

        "first_shift_date": first_shift_date,
        "last_shift_date": last_shift_date,
        "days_since_last_shift": days_since_last_shift,
        "max_consecutive_days": max_consecutive_days,
        "hours_per_week": hours_per_week,
    }


def _build_team_daily(metrics_list: list[dict], date_from: str, date_to: str) -> dict:
    daily = {d: _zero_daily() for d in _date_range(date_from, date_to)}

    for m in metrics_list:
        for d, vals in m.get("daily", {}).items():
            if d not in daily:
                daily[d] = _zero_daily()

            for key, value in vals.items():
                daily[d][key] = daily[d].get(key, 0) + value

    return daily


def _aggregate_dynamics(daily: dict, period_days: int) -> list[dict]:
    period_days = max(int(period_days or 1), 1)

    if period_days <= 60:
        return [
            {"label": d, **daily[d]}
            for d in sorted(daily.keys())
        ]

    groups = {}

    for d in sorted(daily.keys()):
        vals = daily[d]
        dt = _parse_date(d)

        if not dt:
            continue

        if period_days <= 365:
            iso_year, iso_week, _ = dt.isocalendar()
            key = (iso_year, iso_week)
            label = f"{iso_year}-W{iso_week:02d}"
        else:
            key = d[:7]
            label = d[:7]

        if key not in groups:
            groups[key] = {"label": label, **_zero_daily()}

        for field, value in vals.items():
            groups[key][field] += value

    return [groups[key] for key in sorted(groups.keys())]


def _build_employee_insights(m: dict, period_days: int) -> list[str]:
    insights = []

    if m["is_active"] and m["shift_count"] == 0:
        insights.append("Активен, но не было смен за выбранный период.")

    if (
        period_days >= 14
        and m["shift_count"] > 0
        and m["last_shift_date"]
        and m["days_since_last_shift"] is not None
        and m["days_since_last_shift"] > 14
    ):
        insights.append(
            f"Долго не было смен: последняя смена {m['last_shift_date']} "
            f"({m['days_since_last_shift']} дн. назад)."
        )

    if (
        m["taxi_total"] > 0
        and m["total_cost"] > 0
        and m["taxi_share_percent"] > 15
        and m["taxi_total"] >= 1000
    ):
        insights.append(
            f"Высокая доля такси в стоимости: {m['taxi_share_percent']}% "
            f"({m['taxi_total']} ₽)."
        )

    if (
        m["expected_tasks"] > 0
        and m["checklist_percent"] is not None
        and m["checklist_percent"] < 70
    ):
        insights.append(
            f"Низкий процент выполнения чек-листов: {m['checklist_percent']}%."
        )

    if (
        m["shift_count"] > 0
        and m["report_coverage_percent"] is not None
        and m["report_coverage_percent"] < 70
    ):
        insights.append(
            f"Низкое покрытие смен отчётами: {m['report_coverage_percent']}%."
        )

    if m["required_photo_missing"] > 0:
        insights.append(
            f"Есть задачи с обязательным фото без фото: {m['required_photo_missing']} шт."
        )

    if m["max_consecutive_days"] >= 6:
        insights.append(
            f"Возможна перегрузка: {m['max_consecutive_days']} дней со сменами подряд."
        )

    if m["total_hours"] > 0 and m["hours_per_week"] > 45:
        insights.append(
            f"Высокая средняя загрузка: {m['hours_per_week']} ч/нед."
        )

    return insights


def _build_team_insights(metrics_list: list[dict], period_days: int) -> list[str]:
    insights = []

    for m in metrics_list:
        for text in _build_employee_insights(m, period_days):
            insights.append(f"{m['name']}: {text}")

    return insights


# =========================================================
# XLSX SHEET WRITERS
# =========================================================
def _write_sections_sheet(ws, title: str, subtitle: str, sections: list[tuple[str, list[tuple]]]):
    _style_title(ws, 1, 1, title)
    _style_subtitle(ws, 2, 1, subtitle)

    row = 4

    for section_title, items in sections:
        _style_section(ws, row, 1, section_title)
        row += 1

        for label, value in items:
            _style_metric_row(ws, row, label, value, zebra=(row % 2 == 0))
            row += 1

        row += 1

    _auto_width(ws, max_width=80)


def _write_overview_sheet(ws, date_from: str, date_to: str, metrics_list: list[dict]):
    _style_title(ws, 1, 1, "Команда")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    headers = [
        "ФИО",
        "Позиция",
        "Статус",
        "Активен",
        "Ставка ₽/час",
        "Смен",
        "Часов",
        "Заработок ₽",
        "Такси ₽",
        "Полная стоимость ₽",
        "Стоимость часа ₽",
        "Отчётов",
        "% отчётов",
        "Задач",
        "% чек-листов",
        "Задач с фото",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    for m in metrics_list:
        values = [
            m["name"],
            _safe(m["position"]),
            _safe(m["status"]),
            "Да" if m["is_active"] else "Нет",
            _money(m["current_rate"]),
            m["shift_count"],
            _num(m["total_hours"], 1),
            _money(m["earned_total"]),
            _money(m["taxi_total"]),
            _money(m["total_cost"]),
            _money(m["cost_per_hour"]),
            m["reports_count"],
            _safe_percent(m["report_coverage_percent"]),
            m["tasks_completed"],
            _safe_percent(m["checklist_percent"]),
            m["tasks_with_photo"],
        ]

        _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
        row_index += 1

    _auto_width(ws)


def _write_analytics_sheet(ws, date_from: str, date_to: str, metrics_list: list[dict], period_days: int):
    try:
        with get_connection() as conn:
            row = conn.execute("SELECT COUNT(*) AS cnt FROM users WHERE is_active = 0").fetchone()
            hidden_count = row["cnt"] if row else 0
    except Exception:
        hidden_count = 0

    active_count = sum(1 for m in metrics_list if m["is_active"])
    employees_with_shifts = sum(1 for m in metrics_list if m["shift_count"] > 0)

    total_shifts = sum(m["shift_count"] for m in metrics_list)
    total_hours = sum(m["total_hours"] for m in metrics_list)
    total_earned = sum(m["earned_total"] for m in metrics_list)
    total_taxi = sum(m["taxi_total"] for m in metrics_list)
    total_cost = sum(m["total_cost"] for m in metrics_list)

    total_reports = sum(m["reports_count"] for m in metrics_list)
    total_reports_opening = sum(m["reports_opening"] for m in metrics_list)
    total_reports_closing = sum(m["reports_closing"] for m in metrics_list)

    total_shift_dates = sum(m["shift_dates_count"] for m in metrics_list)
    total_covered_report_dates = sum(m["covered_report_dates_count"] for m in metrics_list)

    total_tasks = sum(m["tasks_completed"] for m in metrics_list)
    total_expected_tasks = sum(m["expected_tasks"] for m in metrics_list)
    total_completed_expected = sum(m["completed_expected_tasks"] for m in metrics_list)
    total_tasks_with_photo = sum(m["tasks_with_photo"] for m in metrics_list)
    total_required_photo_missing = sum(m["required_photo_missing"] for m in metrics_list)

    avg_shift_length = _num(total_hours / total_shifts, 1) if total_shifts else 0.0
    avg_hours_per_employee = _num(total_hours / len(metrics_list), 1) if metrics_list else 0.0
    avg_earned_per_employee = _money(total_earned / len(metrics_list)) if metrics_list else 0.0
    avg_taxi_per_employee = _money(total_taxi / len(metrics_list)) if metrics_list else 0.0
    avg_cost_per_employee = _money(total_cost / len(metrics_list)) if metrics_list else 0.0
    avg_cost_per_hour = _money(total_cost / total_hours) if total_hours else 0.0

    report_coverage_total = None
    if total_shift_dates:
        report_coverage_total = _num(total_covered_report_dates / total_shift_dates * 100, 1)

    checklist_total = None
    if total_expected_tasks:
        checklist_total = _num(total_completed_expected / total_expected_tasks * 100, 1)

    photo_percent_total = None
    if total_tasks:
        photo_percent_total = _num(total_tasks_with_photo / total_tasks * 100, 1)

    status_counts = {}
    position_counts = {}

    for m in metrics_list:
        status = m["status"] or "—"
        status_counts[status] = status_counts.get(status, 0) + 1

        position = _normalize_location(m["position"])
        position_counts[position] = position_counts.get(position, 0) + 1

    status_line = ", ".join(f"{k}: {v}" for k, v in sorted(status_counts.items())) or "—"
    position_line = ", ".join(f"{k}: {v}" for k, v in sorted(position_counts.items())) or "—"

    sections = [
        (
            "Период",
            [
                ("Дата начала", date_from),
                ("Дата конца", date_to),
                ("Дней", period_days),
            ],
        ),
        (
            "Команда",
            [
                ("Активных сотрудников", active_count),
                ("Скрытых сотрудников", hidden_count),
                ("Всего в отчёте", len(metrics_list)),
                ("Сотрудников со сменами", employees_with_shifts),
                ("Сотрудников без смен", len(metrics_list) - employees_with_shifts),
                ("Распределение по статусам", status_line),
                ("Распределение по позициям", position_line),
            ],
        ),
        (
            "Работа",
            [
                ("Всего смен", total_shifts),
                ("Всего часов", _num(total_hours, 1)),
                ("Средняя смена, ч", avg_shift_length),
                ("Средняя часовая загрузка на сотрудника", avg_hours_per_employee),
            ],
        ),
        (
            "Финансы",
            [
                ("Всего заработок", _money(total_earned)),
                ("Всего такси", _money(total_taxi)),
                ("Полная стоимость команды", _money(total_cost)),
                ("Средний заработок на сотрудника", avg_earned_per_employee),
                ("Среднее такси на сотрудника", avg_taxi_per_employee),
                ("Средняя полная стоимость на сотрудника", avg_cost_per_employee),
                ("Средняя стоимость часа", avg_cost_per_hour),
            ],
        ),
        (
            "Дисциплина и процессы",
            [
                ("Всего отчётов", total_reports),
                ("Отчётов открытия", total_reports_opening),
                ("Отчётов закрытия", total_reports_closing),
                ("Покрытие смен отчётами", _safe_percent(report_coverage_total)),
                ("Выполнено задач", total_tasks),
                ("Ожидалось задач", total_expected_tasks),
                ("Процент выполнения чек-листов", _safe_percent(checklist_total)),
                ("Задач с фото", total_tasks_with_photo),
                ("Процент задач с фото", _safe_percent(photo_percent_total)),
                ("Обязательных фото без фото", total_required_photo_missing),
            ],
        ),
    ]

    _write_sections_sheet(ws, "Аналитика", f"Период: {date_from} — {date_to}", sections)


def _write_locations_sheet(ws, date_from: str, date_to: str, metrics_list: list[dict]):
    _style_title(ws, 1, 1, "Сводка по локациям")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    loc_metrics = {}

    def ensure(loc: str):
        if loc not in loc_metrics:
            loc_metrics[loc] = {
                "employees": set(),
                "shifts": 0,
                "hours": 0.0,
                "earned": 0.0,
                "taxi": 0.0,
                "tasks": 0,
            }
        return loc_metrics[loc]

    for m in metrics_list:
        position_loc = _normalize_location(m["position"])
        ensure(position_loc)["taxi"] += m["taxi_total"]

        for shift in m["shifts"]:
            loc = _normalize_location(shift.get("location"))
            item = ensure(loc)
            item["employees"].add(m["id"])
            item["shifts"] += 1
            item["hours"] += shift.get("hours") or 0
            item["earned"] += shift.get("earned") or 0

        for act in m["activity"]:
            loc = _normalize_location(act.get("location"))
            ensure(loc)["tasks"] += 1

    headers = [
        "Локация",
        "Сотрудников",
        "Смен",
        "Часов",
        "Заработок ₽",
        "Такси ₽ (по позиции)",
        "Задач выполнено",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    for loc in sorted(loc_metrics.keys()):
        data = loc_metrics[loc]

        values = [
            loc,
            len(data["employees"]),
            data["shifts"],
            _num(data["hours"], 1),
            _money(data["earned"]),
            _money(data["taxi"]),
            data["tasks"],
        ]

        _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
        row_index += 1

    _auto_width(ws)


def _write_dynamics_sheet(ws, date_from: str, date_to: str, dynamics: list[dict]):
    _style_title(ws, 1, 1, "Динамика")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    headers = [
        "Период",
        "Смен",
        "Часов",
        "Заработок ₽",
        "Такси ₽",
        "Задач",
        "Отчётов",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    for item in dynamics:
        values = [
            item["label"],
            item["shifts"],
            _num(item["hours"], 1),
            _money(item["earned"]),
            _money(item["taxi"]),
            item["tasks"],
            item["reports"],
        ]

        _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
        row_index += 1

    _auto_width(ws)


def _write_finance_sheet(ws, date_from: str, date_to: str, metrics_list: list[dict]):
    _style_title(ws, 1, 1, "Финансы")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    headers = [
        "ФИО",
        "Позиция",
        "Текущая ставка ₽/час",
        "Часов",
        "Средневзвешенная ставка ₽/час",
        "Заработок ₽",
        "Такси ₽",
        "Полная стоимость ₽",
        "Стоимость часа ₽",
        "Доля такси",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    for m in metrics_list:
        values = [
            m["name"],
            _safe(m["position"]),
            _money(m["current_rate"]),
            _num(m["total_hours"], 1),
            _money(m["average_rate"]),
            _money(m["earned_total"]),
            _money(m["taxi_total"]),
            _money(m["total_cost"]),
            _money(m["cost_per_hour"]),
            _safe_percent(m["taxi_share_percent"]),
        ]

        _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
        row_index += 1

    _auto_width(ws)


def _write_shifts_sheet(ws, date_from: str, date_to: str, metrics_list: list[dict]):
    _style_title(ws, 1, 1, "Смены")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    headers = [
        "Дата",
        "ФИО",
        "Смена",
        "Локация",
        "Начало",
        "Часов",
        "Ставка ₽/час",
        "Заработок ₽",
        "Активна",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    for m in metrics_list:
        for shift in m["shifts"]:
            values = [
                _safe(shift.get("date")),
                m["name"],
                _safe(shift.get("shift_name")),
                _normalize_location(shift.get("location")),
                _safe(shift.get("start_time")),
                _num(shift.get("hours"), 1),
                _money(shift.get("rate")),
                _money(shift.get("earned")),
                "Да" if shift.get("active") else "Закрыта",
            ]

            _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
            row_index += 1

    _auto_width(ws)


def _write_taxi_sheet(ws, date_from: str, date_to: str, metrics_list: list[dict]):
    _style_title(ws, 1, 1, "Такси")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    headers = [
        "Дата",
        "ФИО",
        "Сумма ₽",
        "Фото",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    for m in metrics_list:
        for expense in m["taxi_expenses"]:
            values = [
                _safe(expense.get("date")),
                m["name"],
                _money(expense.get("amount")),
                "📷 Есть" if expense.get("photos") else "—",
            ]

            _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
            row_index += 1

    _auto_width(ws)


def _write_reports_sheet(ws, date_from: str, date_to: str, metrics_list: list[dict]):
    _style_title(ws, 1, 1, "Отчёты")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    headers = [
        "Дата",
        "ФИО",
        "Тип",
        "Длина текста",
        "Текст",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    for m in metrics_list:
        for report in m["reports"]:
            report_type = "Открытие" if report.get("report_type") == "opening" else "Закрытие"
            full_text = _safe(report.get("full_text"))

            values = [
                _safe(report.get("date")),
                m["name"],
                report_type,
                len(full_text),
                _truncate(full_text, 500),
            ]

            _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
            row_index += 1

    _auto_width(ws, max_width=80)


def _write_checklist_sheet(ws, date_from: str, date_to: str, metrics_list: list[dict]):
    _style_title(ws, 1, 1, "Чек-листы")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    headers = [
        "Дата",
        "ФИО",
        "Задача",
        "Локация",
        "Категория",
        "Время",
        "Фото обязательно",
        "Фото",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    for m in metrics_list:
        for act in m["activity"]:
            values = [
                _safe(act.get("date")),
                m["name"],
                _safe(act.get("item_text")),
                _normalize_location(act.get("location")),
                _safe(act.get("category")),
                _safe(act.get("completed_at")),
                "Да" if act.get("requires_photo") else "—",
                "📷 Есть" if act.get("has_photo") else "—",
            ]

            _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
            row_index += 1

    _auto_width(ws, max_width=60)


def _write_insights_sheet(ws, date_from: str, date_to: str, insights: list[str]):
    _style_title(ws, 1, 1, "Инсайты и риски")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    headers = ["Проблема / рекомендация"]
    _style_header_row(ws, 4, headers)

    row_index = 5

    if not insights:
        _style_data_row(ws, row_index, ["Проблем и рисков не найдено."], zebra=False)
    else:
        for text in insights:
            _style_data_row(ws, row_index, [text], zebra=(row_index % 2 == 0))
            row_index += 1

    _auto_width(ws, max_width=100)


def _write_methodology_sheet(ws, date_from: str, date_to: str):
    _style_title(ws, 1, 1, "Методика расчётов")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    notes = [
        "Часы считаются по плановой длительности смены из типа смены.",
        "Заработок = плановые часы × ставка, действовавшая на дату смены.",
        "Если ставка на дату смены не найдена, используется текущая ставка сотрудника.",
        "Полная стоимость = заработок + расходы на такси.",
        "Стоимость часа = полная стоимость / плановые часы.",
        "Процент отчётов = доля дат со сменами, где есть хотя бы один отчёт.",
        "Процент чек-листов считается относительно ожидаемых задач по локациям и датам смен.",
        "Если ожидаемые задачи не определены, процент чек-листов может быть пустым.",
        "Такси в разрезе локаций распределяется по текущей позиции сотрудника.",
        "В чек-листах учитываются только записи со статусом completed = 1.",
        "Отчёт сформирован автоматически админ-панелью бота.",
    ]

    row_index = 4

    for note in notes:
        cell = ws.cell(row=row_index, column=1, value=note)
        cell.font = Font(name=FONT_NAME, size=10, color=COLOR_DARK)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.row_dimensions[row_index].height = 20
        row_index += 1

    _auto_width(ws, min_width=20, max_width=100)


# =========================================================
# XLSX: ОБЩИЙ ОТЧЁТ ПО ВСЕМ СОТРУДНИКАМ
# =========================================================
def generate_all_employees_report(users: list[dict], period_days: int = 30) -> bytes:
    wb = Workbook()
    date_from, date_to = _period_range(period_days)

    rate_cache = {}
    metrics_list = []

    for user in users:
        try:
            metrics_list.append(
                _employee_metrics(user, date_from, date_to, period_days, rate_cache)
            )
        except Exception as e:
            logger.error("Ошибка расчёта метрик сотрудника %s: %s", user.get("tg_id"), e)

    metrics_list.sort(key=lambda x: (x["name"] or "").lower())

    # 1. Обзор
    ws = wb.active
    ws.title = "Обзор"
    _write_overview_sheet(ws, date_from, date_to, metrics_list)

    # 2. Аналитика
    ws_analytics = wb.create_sheet("Аналитика")
    _write_analytics_sheet(ws_analytics, date_from, date_to, metrics_list, period_days)

    # 3. Локации
    ws_locations = wb.create_sheet("Локации")
    _write_locations_sheet(ws_locations, date_from, date_to, metrics_list)

    # 4. Динамика
    daily = _build_team_daily(metrics_list, date_from, date_to)
    dynamics = _aggregate_dynamics(daily, period_days)

    ws_dynamics = wb.create_sheet("Динамика")
    _write_dynamics_sheet(ws_dynamics, date_from, date_to, dynamics)

    # 5. Финансы
    ws_finance = wb.create_sheet("Финансы")
    _write_finance_sheet(ws_finance, date_from, date_to, metrics_list)

    # 6. Смены
    ws_shifts = wb.create_sheet("Смены")
    _write_shifts_sheet(ws_shifts, date_from, date_to, metrics_list)

    # 7. Такси
    ws_taxi = wb.create_sheet("Такси")
    _write_taxi_sheet(ws_taxi, date_from, date_to, metrics_list)

    # 8. Отчёты
    ws_reports = wb.create_sheet("Отчёты")
    _write_reports_sheet(ws_reports, date_from, date_to, metrics_list)

    # 9. Чек-листы
    ws_check = wb.create_sheet("Чек-листы")
    _write_checklist_sheet(ws_check, date_from, date_to, metrics_list)

    # 10. Инсайты
    insights = _build_team_insights(metrics_list, period_days)
    ws_insights = wb.create_sheet("Инсайты")
    _write_insights_sheet(ws_insights, date_from, date_to, insights)

    # 11. Методика
    ws_methodology = wb.create_sheet("Методика")
    _write_methodology_sheet(ws_methodology, date_from, date_to)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# =========================================================
# XLSX: ОТЧЁТ ПО ОДНОМУ СОТРУДНИКУ
# =========================================================
def _write_employee_profile_sheet(ws, m: dict, date_from: str, date_to: str):
    _style_title(ws, 1, 1, m["name"])
    _style_subtitle(ws, 2, 1, f"Отчёт за период {date_from} — {date_to}")

    rows = [
        ("ФИО", _safe(m["name"])),
        ("Позиция", _safe(m["position"])),
        ("Статус", _safe(m["status"])),
        ("Активен", "Да" if m["is_active"] else "Нет"),
        ("Текущая ставка ₽/час", _money(m["current_rate"])),
        ("Первая смена за период", _safe(m["first_shift_date"])),
        ("Последняя смена за период", _safe(m["last_shift_date"])),
        ("Дней с последней смены", _safe(m["days_since_last_shift"])),
    ]

    row_index = 4

    for label, value in rows:
        _style_metric_row(ws, row_index, label, value, zebra=(row_index % 2 == 0))
        row_index += 1

    _auto_width(ws, max_width=60)


def _write_employee_kpi_sheet(ws, m: dict, date_from: str, date_to: str, period_days: int):
    avg_shift_length = _num(m["total_hours"] / m["shift_count"], 1) if m["shift_count"] else 0.0

    sections = [
        (
            "Работа",
            [
                ("Период", f"{date_from} — {date_to}"),
                ("Дней в периоде", period_days),
                ("Смен", m["shift_count"]),
                ("Часов", _num(m["total_hours"], 1)),
                ("Средняя смена, ч", avg_shift_length),
                ("Первая смена", _safe(m["first_shift_date"])),
                ("Последняя смена", _safe(m["last_shift_date"])),
                ("Дней с последней смены", _safe(m["days_since_last_shift"])),
                ("Максимум смен подряд", m["max_consecutive_days"]),
                ("Средняя загрузка, ч/нед", m["hours_per_week"]),
            ],
        ),
        (
            "Финансы",
            [
                ("Текущая ставка ₽/час", _money(m["current_rate"])),
                ("Средневзвешенная ставка ₽/час", _money(m["average_rate"])),
                ("Заработок ₽", _money(m["earned_total"])),
                ("Такси ₽", _money(m["taxi_total"])),
                ("Полная стоимость ₽", _money(m["total_cost"])),
                ("Стоимость часа ₽", _money(m["cost_per_hour"])),
                ("Доля такси", _safe_percent(m["taxi_share_percent"])),
            ],
        ),
        (
            "Дисциплина и процессы",
            [
                ("Отчётов всего", m["reports_count"]),
                ("Отчётов открытия", m["reports_opening"]),
                ("Отчётов закрытия", m["reports_closing"]),
                ("Дат со сменами", m["shift_dates_count"]),
                ("Дат с отчётами", m["covered_report_dates_count"]),
                ("Покрытие смен отчётами", _safe_percent(m["report_coverage_percent"])),
                ("Выполнено задач", m["tasks_completed"]),
                ("Ожидалось задач", m["expected_tasks"]),
                ("Процент выполнения чек-листов", _safe_percent(m["checklist_percent"])),
                ("Задач с фото", m["tasks_with_photo"]),
                ("Процент задач с фото", _safe_percent(m["photo_percent"])),
                ("Обязательных фото без фото", m["required_photo_missing"]),
            ],
        ),
    ]

    _write_sections_sheet(ws, "KPI", f"Сотрудник: {m['name']}", sections)


def _write_employee_finance_sheet(ws, m: dict, date_from: str, date_to: str):
    _style_title(ws, 1, 1, "Заработок по дням")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    daily = {}

    for shift in m["shifts"]:
        d = shift.get("date")

        if not d:
            continue

        if d not in daily:
            daily[d] = {
                "shifts": 0,
                "hours": 0.0,
                "earned": 0.0,
                "taxi": 0.0,
            }

        daily[d]["shifts"] += 1
        daily[d]["hours"] += shift.get("hours") or 0
        daily[d]["earned"] += shift.get("earned") or 0

    for expense in m["taxi_expenses"]:
        d = expense.get("date")

        if not d:
            continue

        if d not in daily:
            daily[d] = {
                "shifts": 0,
                "hours": 0.0,
                "earned": 0.0,
                "taxi": 0.0,
            }

        daily[d]["taxi"] += _money(expense.get("amount"))

    headers = [
        "Дата",
        "Смен",
        "Часов",
        "Заработок ₽",
        "Такси ₽",
        "Итого ₽",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    if not daily:
        _style_data_row(ws, row_index, ["Нет данных за период"], zebra=False)
    else:
        for d in sorted(daily.keys()):
            item = daily[d]
            total = _money(item["earned"] + item["taxi"])

            values = [
                d,
                item["shifts"],
                _num(item["hours"], 1),
                _money(item["earned"]),
                _money(item["taxi"]),
                total,
            ]

            _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
            row_index += 1

    _auto_width(ws)


def _write_employee_shifts_sheet(ws, m: dict, date_from: str, date_to: str):
    _style_title(ws, 1, 1, "Смены")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    headers = [
        "Дата",
        "Смена",
        "Локация",
        "Начало",
        "Часов",
        "Ставка ₽/час",
        "Заработок ₽",
        "Активна",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    if not m["shifts"]:
        _style_data_row(ws, row_index, ["Смен за период нет"], zebra=False)
    else:
        for shift in m["shifts"]:
            values = [
                _safe(shift.get("date")),
                _safe(shift.get("shift_name")),
                _normalize_location(shift.get("location")),
                _safe(shift.get("start_time")),
                _num(shift.get("hours"), 1),
                _money(shift.get("rate")),
                _money(shift.get("earned")),
                "Да" if shift.get("active") else "Закрыта",
            ]

            _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
            row_index += 1

    _auto_width(ws)


def _write_employee_taxi_sheet(ws, m: dict, date_from: str, date_to: str):
    _style_title(ws, 1, 1, "Такси")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    headers = [
        "Дата",
        "Сумма ₽",
        "Фото",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    if not m["taxi_expenses"]:
        _style_data_row(ws, row_index, ["Поездок за период нет"], zebra=False)
    else:
        for expense in m["taxi_expenses"]:
            values = [
                _safe(expense.get("date")),
                _money(expense.get("amount")),
                "📷 Есть" if expense.get("photos") else "—",
            ]

            _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
            row_index += 1

    _auto_width(ws)


def _write_employee_reports_sheet(ws, m: dict, date_from: str, date_to: str):
    _style_title(ws, 1, 1, "Отчёты")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    headers = [
        "Дата",
        "Тип",
        "Длина текста",
        "Текст",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    if not m["reports"]:
        _style_data_row(ws, row_index, ["Отчётов за период нет"], zebra=False)
    else:
        for report in m["reports"]:
            report_type = "Открытие" if report.get("report_type") == "opening" else "Закрытие"
            full_text = _safe(report.get("full_text"))

            values = [
                _safe(report.get("date")),
                report_type,
                len(full_text),
                _truncate(full_text, 700),
            ]

            _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
            row_index += 1

    _auto_width(ws, max_width=80)


def _write_employee_checklist_sheet(ws, m: dict, date_from: str, date_to: str):
    _style_title(ws, 1, 1, "Чек-листы")
    _style_subtitle(ws, 2, 1, f"Период: {date_from} — {date_to}")

    headers = [
        "Дата",
        "Задача",
        "Локация",
        "Категория",
        "Время",
        "Фото обязательно",
        "Фото",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    if not m["activity"]:
        _style_data_row(ws, row_index, ["Выполненных задач за период нет"], zebra=False)
    else:
        for act in m["activity"]:
            values = [
                _safe(act.get("date")),
                _safe(act.get("item_text")),
                _normalize_location(act.get("location")),
                _safe(act.get("category")),
                _safe(act.get("completed_at")),
                "Да" if act.get("requires_photo") else "—",
                "📷 Есть" if act.get("has_photo") else "—",
            ]

            _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
            row_index += 1

    _auto_width(ws, max_width=60)


def _write_employee_rates_sheet(ws, user_id: int):
    rates = get_salary_history(user_id)

    _style_title(ws, 1, 1, "История ставок")
    _style_subtitle(ws, 2, 1, "Все изменения ставки")

    headers = [
        "Дата от",
        "Дата до",
        "Ставка ₽/час",
    ]

    _style_header_row(ws, 4, headers)

    row_index = 5

    if not rates:
        _style_data_row(ws, row_index, ["История ставок пуста"], zebra=False)
    else:
        for rate in rates:
            values = [
                _safe(rate.get("date_from")),
                _safe(rate.get("date_to")) or "Активна",
                _money(rate.get("rate")),
            ]

            _style_data_row(ws, row_index, values, zebra=(row_index % 2 == 0))
            row_index += 1

    _auto_width(ws)


def generate_employee_report(user_id: int, period_days: int = 30) -> bytes:
    wb = Workbook()

    user = get_user(user_id) or {"tg_id": user_id}
    date_from, date_to = _period_range(period_days)

    rate_cache = {}
    m = _employee_metrics(user, date_from, date_to, period_days, rate_cache)

    # 1. Профиль
    ws = wb.active
    ws.title = "Профиль"
    _write_employee_profile_sheet(ws, m, date_from, date_to)

    # 2. KPI
    ws_kpi = wb.create_sheet("KPI")
    _write_employee_kpi_sheet(ws_kpi, m, date_from, date_to, period_days)

    # 3. Заработок
    ws_finance = wb.create_sheet("Заработок")
    _write_employee_finance_sheet(ws_finance, m, date_from, date_to)

    # 4. Смены
    ws_shifts = wb.create_sheet("Смены")
    _write_employee_shifts_sheet(ws_shifts, m, date_from, date_to)

    # 5. Такси
    ws_taxi = wb.create_sheet("Такси")
    _write_employee_taxi_sheet(ws_taxi, m, date_from, date_to)

    # 6. Отчёты
    ws_reports = wb.create_sheet("Отчёты")
    _write_employee_reports_sheet(ws_reports, m, date_from, date_to)

    # 7. Чек-листы
    ws_check = wb.create_sheet("Чек-листы")
    _write_employee_checklist_sheet(ws_check, m, date_from, date_to)

    # 8. Ставки
    ws_rates = wb.create_sheet("Ставки")
    _write_employee_rates_sheet(ws_rates, user_id)

    # 9. Динамика
    daily = {d: _zero_daily() for d in _date_range(date_from, date_to)}

    for d, vals in m.get("daily", {}).items():
        if d not in daily:
            daily[d] = _zero_daily()

        for key, value in vals.items():
            daily[d][key] += value

    dynamics = _aggregate_dynamics(daily, period_days)

    ws_dynamics = wb.create_sheet("Динамика")
    _write_dynamics_sheet(ws_dynamics, date_from, date_to, dynamics)

    # 10. Инсайты
    insights = _build_employee_insights(m, period_days)
    ws_insights = wb.create_sheet("Инсайты")
    _write_insights_sheet(ws_insights, date_from, date_to, insights)

    # 11. Методика
    ws_methodology = wb.create_sheet("Методика")
    _write_methodology_sheet(ws_methodology, date_from, date_to)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# =========================================================
# TAXI PHOTOS COLLECTOR
# =========================================================
def collect_taxi_photo_ids(user_id: int, period_days: int = 30, limit: int = 30) -> list[str]:
    date_from, date_to = _period_range(period_days)
    expenses = get_taxi_expenses_full(user_id, date_from, date_to)

    photo_ids: list[str] = []

    for expense in expenses:
        for photo_id in expense.get("photos", []):
            photo_ids.append(photo_id)

            if len(photo_ids) >= limit:
                return photo_ids

    return photo_ids


# =========================================================
# УДАЛЕНИЕ СОТРУДНИКА
# =========================================================
def delete_employee_completely(tg_id: int) -> None:
    """
    Полное удаление сотрудника со всей историей:
    - смены
    - такси
    - отчёты
    - прогресс чек-листов (как completed_by)
    - ставки
    - профиль (таблица users)
    """
    with get_connection() as conn:
        conn.execute("DELETE FROM shifts WHERE user_id = ?", (tg_id,))
        logger.info("Удалены смены для %s", tg_id)

        conn.execute("DELETE FROM taxi_expenses WHERE user_id = ?", (tg_id,))
        logger.info("Удалены такси для %s", tg_id)

        conn.execute("DELETE FROM shift_reports WHERE author_id = ?", (tg_id,))
        logger.info("Удалены отчёты для %s", tg_id)

        conn.execute(
            "UPDATE checklist_shared_progress SET completed_by = NULL WHERE completed_by = ?",
            (tg_id,)
        )
        logger.info("Обнулены completed_by для %s", tg_id)

        conn.execute("DELETE FROM salary_rates WHERE user_id = ?", (tg_id,))
        logger.info("Удалены ставки для %s", tg_id)

        conn.execute("DELETE FROM users WHERE tg_id = ?", (tg_id,))
        logger.info("Удалён пользователь %s", tg_id)

        conn.commit()